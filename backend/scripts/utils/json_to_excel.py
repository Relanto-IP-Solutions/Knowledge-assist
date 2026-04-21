"""Convert agent output JSON (e.g. data/output/agent_OPP-1011.json) to Excel.

Uses data from "final_answers" only. final_answers is a dict keyed by question_id;
each value has answer, confidence, sources, etc. One row per question.
The "question" and "answer_type" columns are filled from the sase_questions table when DB is available.

Usage:
  uv run python scripts/utils/json_to_excel.py data/output/agent_OPP-1011.json
  uv run python scripts/utils/json_to_excel.py data/output/agent_OPP-1011.json -o data/output/agent_OPP-1011.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


_PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_PROJECT_ROOT))

# Load env so DB connection works for sase_questions
try:
    from dotenv import load_dotenv

    for _ef in (
        _PROJECT_ROOT / "configs" / ".env",
        _PROJECT_ROOT / "configs" / "secrets" / ".env",
    ):
        if _ef.exists():
            load_dotenv(_ef, override=False)
except ImportError:
    pass

# Preferred column order for final_answers (question_id, question, answer_type, then answer, confidence, rest)
_PREFERRED_ORDER = ("question_id", "question", "answer_type", "answer", "confidence")


def _column_order(all_keys: set[str]) -> list[str]:
    """Return column headers: preferred first (if present), then rest sorted."""
    preferred = [k for k in _PREFERRED_ORDER if k in all_keys]
    rest = sorted(all_keys - set(_PREFERRED_ORDER))
    return preferred + rest


def _cell_value(val) -> str:
    """Serialize value for Excel cell: dict/list -> pretty-printed JSON to retain structure."""
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False, indent=2)
    return str(val)


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert agent JSON to Excel")
    parser.add_argument("json_path", type=Path, help="Path to agent_OPP-*.json")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output Excel path (default: same name with .xlsx)",
    )
    args = parser.parse_args()

    json_path = (
        args.json_path
        if args.json_path.is_absolute()
        else _PROJECT_ROOT / args.json_path
    )
    if not json_path.exists():
        print(f"File not found: {json_path}", file=sys.stderr)
        return 1

    out_path = args.output
    if out_path is None:
        out_path = json_path.with_suffix(".xlsx")
    elif not out_path.is_absolute():
        out_path = _PROJECT_ROOT / out_path

    try:
        import openpyxl
    except ImportError:
        print(
            "Install openpyxl: uv add openpyxl  or  pip install openpyxl",
            file=sys.stderr,
        )
        return 1

    data = json.loads(json_path.read_text(encoding="utf-8"))
    final_answers = data.get("final_answers", {})
    if not final_answers or not isinstance(final_answers, dict):
        print("No 'final_answers' dict in JSON.", file=sys.stderr)
        return 1

    # Convert to list of rows: [{"question_id": qid, **val}, ...]
    rows = []
    for qid, val in final_answers.items():
        if not isinstance(val, dict):
            rows.append({"question_id": qid})
        else:
            rows.append({"question_id": qid, **val})

    # Populate question and answer_type from sase_questions (q_id -> question, answer_type)
    questions_and_types: dict[str, dict[str, str]] = {}
    try:
        from src.services.rag_engine.retrieval import load_questions_and_answer_types

        questions_and_types = load_questions_and_answer_types()
    except Exception as e:
        print(
            f"Could not load sase_questions (question and answer_type columns will be empty): {e}",
            file=sys.stderr,
        )
    for row in rows:
        qid = row["question_id"]
        info = questions_and_types.get(qid, {})
        row["question"] = info.get("question", "")
        row["answer_type"] = info.get("answer_type", "")

    all_keys = set()
    for entry in rows:
        all_keys.update(entry.keys())
    headers = _column_order(all_keys)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FinalAnswers"

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, entry in enumerate(rows, start=2):
        for col, key in enumerate(headers, 1):
            val = entry.get(key)
            cell_val = _cell_value(val)
            cell = ws.cell(row=row_idx, column=col, value=cell_val)
            # Wrap text for dict/list columns so pretty-printed JSON is readable
            if isinstance(val, (dict, list)):
                cell.alignment = openpyxl.styles.Alignment(
                    wrap_text=True, vertical="top"
                )

    # Default column widths (A=question_id, B=question, C=answer_type, D=answer, E=confidence; rest 60)
    widths = {"A": 16, "B": 50, "C": 18, "D": 50, "E": 12}
    for i, _key in enumerate(headers):
        col_letter = openpyxl.utils.get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = widths.get(col_letter, 60)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written: {out_path} ({len(headers)} columns, {len(rows)} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
