"""GCS raw pipeline: read recent raw source files from GCS, preprocess by type, write to processed tier.

Sources and their handlers:
    zoom      → VTTPreprocessor           → {opp_id}/processed/zoom_transcripts/{stem}.txt
    slack     → SlackOrchestrator         → {opp_id}/processed/slack_messages/{channel_id}/summary.txt
    documents → DocumentExtractionService → {opp_id}/processed/documents/{stem}.txt
    onedrive  → DocumentExtractionService → {opp_id}/processed/onedrive/{stem}.txt
    gmail     → GmailPreprocessor         → {opp_id}/processed/gmail_messages/{thread_id}/content.txt

GCS path conventions:

    Zoom:
        Raw input      : {opp_id}/raw/zoom/{meeting_id}.vtt
        Processed text : {opp_id}/processed/zoom_transcripts/{meeting_id}.txt

    Slack:
        Raw messages   : {opp_id}/raw/slack/{channel_id}/slack_messages.json
        Metadata       : {opp_id}/raw/slack/slack_metadata.json   (required)
        Processed text : {opp_id}/processed/slack_messages/{channel_id}/summary.txt   (RAG-ready)
        State file     : {opp_id}/processed/slack_messages/{channel_id}/state.json   (checkpoint)

    Gmail:
        Raw thread     : {opp_id}/raw/gmail/{thread_id}/thread.json
        Processed text : {opp_id}/processed/gmail_messages/{thread_id}/content.txt   (RAG-ready)
        State file     : {opp_id}/processed/gmail_messages/{thread_id}/state.json    (checkpoint)

The pipeline is designed to be idempotent: for Slack files the state file's
`last_processed_ts` is used to skip messages already seen, so reprocessing
an unchanged file is a no-op. A `since` parameter allows the caller to
restrict processing to blobs uploaded within a rolling time window, which is
used by the 15-minute Cloud Scheduler trigger in functions/gcs_file_processor.py.
"""

import json
import re
from datetime import datetime
from pathlib import PurePosixPath

from configs.settings import get_settings
from src.services.document_extraction import DocumentExtractionService
from src.services.document_extraction.extractors.spreadsheet import (
    SpreadsheetExtractor,
    processed_xlsx_sheet_filename,
)
from src.services.preprocessing.mail import GmailPreprocessor
from src.services.preprocessing.slack import ChannelAnalysis, SlackOrchestrator
from src.services.preprocessing.slack.formatter import SlackAnalysisFormatter
from src.services.preprocessing.zoom import VTTPreprocessor
from src.services.storage import Storage
from src.utils.logger import get_logger


logger = get_logger(__name__)

# Pattern that raw blob paths must match: {opp_id}/raw/{source}/{object_name}
_RAW_BLOB_RE = re.compile(r"^([^/]+)/raw/([^/]+)/(.+)$")

# Document extraction: supported extensions for raw/documents
_DOCUMENT_EXTENSIONS = frozenset({
    ".pdf",
    ".docx",
    ".md",
    ".markdown",
    ".pptx",
    ".xlsx",
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
})

# Max size for raw documents (bytes). Files larger than this are skipped during raw→processed.
MAX_DOCUMENT_SIZE_BYTES = 50 * 1024 * 1024


class GcsPipeline:
    """Preprocesses raw source files from GCS by type and writes to processed tier."""

    def __init__(self, storage: Storage | None = None) -> None:
        self._storage = storage or Storage()
        self._formatter = SlackAnalysisFormatter()
        self._document_extractor = DocumentExtractionService()
        self._spreadsheet_extractor = SpreadsheetExtractor()
        self._gmail_preprocessor = GmailPreprocessor()

    def _process_zoom(self, opp_id: str, object_name: str) -> str | None:
        """Preprocess a Zoom VTT transcript and write to processed/zoom_transcripts/.

        Returns:
            GCS URI of the written processed object, or None if skipped.
        """
        if not object_name.lower().endswith(".vtt"):
            logger.bind(opportunity_id=opp_id).debug(
                "Skipping non-VTT file in zoom source"
            )
            return None

        dest_name = PurePosixPath(object_name).stem + ".txt"

        raw_updated_at = self._storage.blob_updated_at(
            "raw", opp_id, "zoom", object_name
        )
        processed_updated_at = self._storage.blob_updated_at(
            "processed", opp_id, "zoom_transcripts", dest_name
        )

        if (
            processed_updated_at is not None
            and raw_updated_at is not None
            and processed_updated_at >= raw_updated_at
        ):
            logger.bind(opportunity_id=opp_id).info(
                "Skipping transcript — processed output is up to date"
            )
            return None

        logger.bind(opportunity_id=opp_id).info("Processing Zoom VTT transcript")
        raw_bytes = self._storage.read("raw", opp_id, "zoom", object_name)
        df = VTTPreprocessor().preprocess(raw_bytes)
        tsv_content = df.to_csv(sep="\t", index=False)

        uri = self._storage.write(
            tier="processed",
            opportunity_id=opp_id,
            source="zoom_transcripts",
            object_name=dest_name,
            content=tsv_content,
            content_type="text/plain",
        )
        logger.bind(opportunity_id=opp_id).info("Processed VTT transcript")
        return uri

    def _process_slack(self, opp_id: str, object_name: str) -> str | None:
        """Preprocess a Slack message export and write analysis outputs to GCS.

        Expected raw path: raw/slack/{channel_id}/slack_messages.json
        Returns GCS URI of the written *_summary.txt object, or None if skipped.
        """
        if object_name == "slack_metadata.json":
            logger.bind(opportunity_id=opp_id).debug("Skipping Slack metadata file")
            return None

        if not object_name.endswith("/slack_messages.json"):
            logger.bind(opportunity_id=opp_id).debug("Skipping non-message Slack file")
            return None

        channel = PurePosixPath(object_name).parts[0]

        user_map: dict[str, str] | None = None
        try:
            meta_bytes = self._storage.read(
                "raw", opp_id, "slack", "slack_metadata.json"
            )
            metadata = json.loads(meta_bytes)
            user_map = {
                member["id"]: member["name"]
                for ch in metadata.get("channels", [])
                for member in ch.get("members", [])
                if member.get("id") and member.get("name")
            }
            logger.bind(opportunity_id=opp_id).debug("Loaded Slack user map")
        except FileNotFoundError:
            logger.bind(opportunity_id=opp_id).warning(
                "slack_metadata.json not found — user IDs will not be resolved",
                exc_info=True,
            )

        since_ts: float | None = None
        previous_analysis: ChannelAnalysis | None = None
        state_object = f"{channel}/state.json"
        try:
            state = json.loads(
                self._storage.read("processed", opp_id, "slack_messages", state_object)
            )
            since_ts = state.get("last_processed_ts")
            if raw_analysis := state.get("analysis"):
                previous_analysis = ChannelAnalysis.model_validate(raw_analysis)
            logger.bind(opportunity_id=opp_id).debug("Loaded Slack channel state")
        except FileNotFoundError:
            pass

        logger.bind(opportunity_id=opp_id).info("Processing Slack channel")
        raw_bytes = self._storage.read("raw", opp_id, "slack", object_name)

        result = SlackOrchestrator().process(
            raw_bytes=raw_bytes,
            channel=channel,
            opportunity_id=opp_id,
            since_ts=since_ts,
            user_map=user_map,
            previous_analysis=previous_analysis,
        )

        if result is None:
            logger.bind(opportunity_id=opp_id).info(
                "No new Slack messages — skipping GCS writes"
            )
            return None

        analysis, latest_ts = result

        summary_txt = self._formatter.format_analysis_as_text(
            analysis, channel, opp_id, latest_ts
        )
        txt_uri = self._storage.write(
            tier="processed",
            opportunity_id=opp_id,
            source="slack_messages",
            object_name=f"{channel}/summary.txt",
            content=summary_txt,
            content_type="text/plain",
        )

        self._storage.write(
            tier="processed",
            opportunity_id=opp_id,
            source="slack_messages",
            object_name=state_object,
            content=json.dumps({
                "last_processed_ts": latest_ts,
                "analysis": analysis.model_dump(),
            }),
            content_type="application/json",
        )

        logger.bind(opportunity_id=opp_id).info("Processed Slack channel")
        return txt_uri

    def _process_documents_from_source(
        self,
        opp_id: str,
        object_name: str,
        *,
        raw_source: str,
        processed_source: str,
    ) -> list[str] | None:
        """Extract text from document and write to processed/{processed_source}.

        Returns:
            List of GCS URIs written, or None if skipped.
        """
        ext = (
            "." + (object_name or "").lower().rsplit(".", 1)[-1]
            if "." in (object_name or "")
            else ""
        )
        if ext not in _DOCUMENT_EXTENSIONS:
            logger.bind(opportunity_id=opp_id, object_name=object_name).debug(
                "Skipping unsupported document type"
            )
            return None

        raw_stem = PurePosixPath(object_name).stem
        dest_name = raw_stem + ".txt"

        raw_updated_at = self._storage.blob_updated_at(
            "raw", opp_id, raw_source, object_name
        )
        processed_updated_at = self._storage.blob_updated_at(
            "processed", opp_id, processed_source, dest_name
        )

        # XLSX: write one processed TXT per sheet (local-compatible naming).
        if ext == ".xlsx":
            # Determine if all expected per-sheet outputs are up to date.
            # We load the workbook metadata to get sheet names; this is cheap relative
            # to full extraction and avoids rewriting unchanged files.
            try:
                from io import BytesIO

                from openpyxl import load_workbook

                raw_bytes_probe = self._storage.read(
                    "raw", opp_id, raw_source, object_name
                )
                wb = load_workbook(
                    BytesIO(raw_bytes_probe), data_only=True, read_only=True
                )
                expected_names = [
                    processed_xlsx_sheet_filename(original_stem=raw_stem, sheet_name=s)
                    for s in wb.sheetnames
                ]
                wb.close()
            except Exception:
                expected_names = []

            if raw_updated_at is not None and expected_names:
                all_current = True
                for name in expected_names:
                    ts = self._storage.blob_updated_at(
                        "processed", opp_id, processed_source, name
                    )
                    if ts is None or ts < raw_updated_at:
                        all_current = False
                        break
                if all_current:
                    logger.bind(opportunity_id=opp_id, object_name=object_name).info(
                        "Skipping spreadsheet — all processed sheets are up to date"
                    )
                    return None

            raw_size = self._storage.blob_size("raw", opp_id, raw_source, object_name)
            if raw_size is not None and raw_size > MAX_DOCUMENT_SIZE_BYTES:
                max_size_mb = int(MAX_DOCUMENT_SIZE_BYTES / (1024 * 1024))
                logger.bind(opportunity_id=opp_id).warning(
                    "Spreadsheet %s exceeds %s MB size limit. Skipping raw→processed.",
                    object_name,
                    max_size_mb,
                )
                return None

            logger.bind(opportunity_id=opp_id, object_name=object_name).info(
                "Processing spreadsheet (.xlsx)"
            )
            raw_bytes = self._storage.read("raw", opp_id, raw_source, object_name)
            extracted = self._spreadsheet_extractor.extract_sheets(
                raw_bytes, object_name
            )
            written: list[str] = []
            for sheet in extracted:
                uri = self._storage.write(
                    tier="processed",
                    opportunity_id=opp_id,
                    source=processed_source,
                    object_name=sheet.processed_object_name,
                    content=sheet.text,
                    content_type="text/plain",
                )
                written.append(uri)
            logger.bind(
                opportunity_id=opp_id,
                object_name=object_name,
                sheets_written=len(written),
            ).info("Processed spreadsheet")
            return written

        raw_size = self._storage.blob_size("raw", opp_id, raw_source, object_name)
        if raw_size is not None and raw_size > MAX_DOCUMENT_SIZE_BYTES:
            max_size_mb = int(MAX_DOCUMENT_SIZE_BYTES / (1024 * 1024))
            logger.bind(opportunity_id=opp_id).warning(
                "Document %s exceeds %s MB size limit. Skipping raw→processed.",
                object_name,
                max_size_mb,
            )
            return None

        if (
            processed_updated_at is not None
            and raw_updated_at is not None
            and processed_updated_at >= raw_updated_at
        ):
            logger.bind(opportunity_id=opp_id).info(
                "Skipping document — processed output is up to date"
            )
            return None

        logger.bind(opportunity_id=opp_id, object_name=object_name).info(
            "Processing document"
        )
        raw_bytes = self._storage.read("raw", opp_id, raw_source, object_name)

        try:
            extracted_text = self._document_extractor.extract(raw_bytes, object_name)
        except ValueError as exc:
            logger.bind(opportunity_id=opp_id, object_name=object_name).warning(
                "Document extraction skipped: %s",
                exc,
            )
            return None

        uri = self._storage.write(
            tier="processed",
            opportunity_id=opp_id,
            source=processed_source,
            object_name=dest_name,
            content=extracted_text,
            content_type="text/plain",
        )
        logger.bind(opportunity_id=opp_id, object_name=object_name).info(
            "Processed document"
        )
        return [uri]

    def _process_documents(self, opp_id: str, object_name: str) -> list[str] | None:
        return self._process_documents_from_source(
            opp_id,
            object_name,
            raw_source="documents",
            processed_source="documents",
        )

    def _process_onedrive_documents(self, opp_id: str, object_name: str) -> list[str] | None:
        return self._process_documents_from_source(
            opp_id,
            object_name,
            raw_source="onedrive",
            processed_source="onedrive",
        )

    def _process_gmail(self, opp_id: str, object_name: str) -> str | None:
        """Preprocess a Gmail thread JSON and write to processed/gmail_messages/.

        Expected raw path: raw/gmail/{thread_id}/thread.json
        Returns GCS URI of the written content.txt object, or None if skipped.
        """
        # Only process thread.json files
        if not object_name.endswith("/thread.json"):
            logger.bind(opportunity_id=opp_id).debug(
                "Skipping non-thread Gmail file: %s", object_name
            )
            return None

        # Extract thread_id from path (e.g., "abc123/thread.json" -> "abc123")
        thread_id = PurePosixPath(object_name).parent.name
        if not thread_id or thread_id == ".":
            # Handle flat path like "thread.json" (shouldn't happen with new connector)
            thread_id = PurePosixPath(object_name).stem

        # Check if processed output is up to date
        raw_updated_at = self._storage.blob_updated_at(
            "raw", opp_id, "gmail", object_name
        )
        processed_path = f"{thread_id}/content.txt"
        processed_updated_at = self._storage.blob_updated_at(
            "processed", opp_id, "gmail_messages", processed_path
        )

        if (
            processed_updated_at is not None
            and raw_updated_at is not None
            and processed_updated_at >= raw_updated_at
        ):
            logger.bind(opportunity_id=opp_id).info(
                "Skipping Gmail thread — processed output is up to date: %s",
                thread_id,
            )
            return None

        logger.bind(opportunity_id=opp_id).info(
            "Processing Gmail thread: %s", thread_id
        )

        # Read and preprocess
        raw_bytes = self._storage.read("raw", opp_id, "gmail", object_name)
        try:
            cleaned_thread, formatted_text = (
                self._gmail_preprocessor.preprocess_and_format(raw_bytes)
            )
        except ValueError as exc:
            logger.bind(opportunity_id=opp_id).warning(
                "Gmail preprocessing failed for %s: %s",
                thread_id,
                exc,
            )
            return None

        # Skip if no content after cleaning
        if cleaned_thread.content_message_count == 0:
            logger.bind(opportunity_id=opp_id).info(
                "Gmail thread has no content after cleaning: %s", thread_id
            )
            return None

        # Write processed content
        txt_uri = self._storage.write(
            tier="processed",
            opportunity_id=opp_id,
            source="gmail_messages",
            object_name=processed_path,
            content=formatted_text,
            content_type="text/plain",
        )

        # Write state file for tracking
        state_path = f"{thread_id}/state.json"
        last_message = cleaned_thread.messages[-1] if cleaned_thread.messages else None
        self._storage.write(
            tier="processed",
            opportunity_id=opp_id,
            source="gmail_messages",
            object_name=state_path,
            content=json.dumps({
                "thread_id": thread_id,
                "last_message_id": last_message.id if last_message else None,
                "message_count": cleaned_thread.message_count,
                "content_message_count": cleaned_thread.content_message_count,
                "processed_at": datetime.utcnow().isoformat(),
            }),
            content_type="application/json",
        )

        logger.bind(opportunity_id=opp_id).info(
            "Processed Gmail thread: %s (%d/%d messages with content)",
            thread_id,
            cleaned_thread.content_message_count,
            cleaned_thread.message_count,
        )
        return txt_uri

    def _get_handler(self, source: str):
        """Return the handler method for the given source name."""
        handlers = {
            "zoom": self._process_zoom,
            "slack": self._process_slack,
            "documents": self._process_documents,
            "onedrive": self._process_onedrive_documents,
            "gmail": self._process_gmail,
        }
        return handlers.get(source)

    def _reconcile_documents_orphans(self, opportunity_id: str) -> list[str]:
        """Delete processed/documents files whose raw counterpart no longer exists.

        When a raw document is removed, its processed .txt output becomes an orphan.
        This method detects such orphans and deletes them from processed/documents.

        Returns:
            List of GCS URIs that were deleted.
        """
        deleted_uris: list[str] = []
        try:
            raw_objects = self._storage.list_objects("raw", opportunity_id, "documents")
            raw_stems: set[str] = set()
            raw_xlsx_stems: set[str] = set()
            for obj in raw_objects:
                ext = (
                    "." + (obj or "").lower().rsplit(".", 1)[-1]
                    if "." in (obj or "")
                    else ""
                )
                if ext in _DOCUMENT_EXTENSIONS:
                    stem = PurePosixPath(obj).stem
                    raw_stems.add(stem)
                    if ext == ".xlsx":
                        raw_xlsx_stems.add(stem)

            processed_objects = self._storage.list_objects(
                "processed", opportunity_id, "documents"
            )
            for obj in processed_objects:
                # Normal processed documents are {stem}.txt, so stem matching works.
                # Spreadsheet processed outputs are per-sheet and follow:
                #   _pre_xlsx_{sheet}_{raw_stem}.txt
                # We treat those as orphans if the raw XLSX stem no longer exists.
                obj_name = PurePosixPath(obj).name
                stem = PurePosixPath(obj).stem

                is_sheet_output = obj_name.startswith(
                    "_pre_xlsx_"
                ) and obj_name.endswith(".txt")
                if is_sheet_output:
                    # If this matches any raw XLSX stem suffix, keep it; else delete.
                    # Generic rule: delete if there is no raw XLSX stem such that
                    # the processed filename ends with _{stem}.txt.
                    keep = False
                    for xstem in raw_xlsx_stems:
                        if obj_name.endswith(f"_{xstem}.txt"):
                            keep = True
                            break
                    if keep:
                        continue

                    self._storage.delete("processed", opportunity_id, "documents", obj)
                    bucket_name = get_settings().ingestion.gcs_bucket_ingestion
                    path = f"{opportunity_id}/processed/documents/{obj}"
                    uri = f"gs://{bucket_name}/{path}"
                    deleted_uris.append(uri)
                    logger.bind(opportunity_id=opportunity_id).info(
                        "Reconciliation: deleted orphan spreadsheet sheet output — object_name=%s",
                        obj,
                    )
                    continue

                if stem not in raw_stems:
                    self._storage.delete("processed", opportunity_id, "documents", obj)
                    bucket_name = get_settings().ingestion.gcs_bucket_ingestion
                    path = f"{opportunity_id}/processed/documents/{obj}"
                    uri = f"gs://{bucket_name}/{path}"
                    deleted_uris.append(uri)
                    logger.bind(opportunity_id=opportunity_id).info(
                        "Reconciliation: deleted orphan processed document — object_name=%s",
                        obj,
                    )
        except Exception:
            logger.bind(opportunity_id=opportunity_id).exception(
                "Reconciliation: failed to reconcile documents orphans — opportunity_id=%s",
                opportunity_id,
            )
        return deleted_uris

    def run_opportunity(
        self,
        opportunity_id: str,
        since: datetime | None = None,
    ) -> tuple[list[str], list[str]]:
        """Preprocess all raw source files for a single opportunity.

        Runs documents orphan reconciliation first (delete processed files whose
        raw counterpart was removed), then processes raw files.

        Returns:
            Tuple of (written_uris, deleted_uris).
        """
        deleted = self._reconcile_documents_orphans(opportunity_id)
        written: list[str] = []
        for source in ("zoom", "slack", "documents", "onedrive", "gmail"):
            handler = self._get_handler(source)
            if handler is None:
                continue
            prefix = f"{opportunity_id}/raw/{source}/"
            blob_names = self._storage.list_blobs_since(prefix=prefix, since=since)

            for full_path in blob_names:
                object_name = full_path[len(prefix) :]
                if not object_name:
                    continue
                try:
                    result = handler(opportunity_id, object_name)
                    if not result:
                        continue
                    # Most handlers return a single URI string; documents may return multiple.
                    if isinstance(result, list):
                        written.extend(result)
                    else:
                        written.append(result)
                except NotImplementedError:
                    logger.bind(opportunity_id=opportunity_id).debug(
                        "Source handler not yet implemented, skipping"
                    )
                except Exception:
                    logger.bind(opportunity_id=opportunity_id).error(
                        "Failed to process file",
                        exc_info=True,
                    )
                    raise

        return written, deleted

    def run(
        self,
        opportunity_id: str | None = None,
        since: datetime | None = None,
    ) -> tuple[list[str], list[str]]:
        """Preprocess raw source files across one or all opportunities.

        When opportunity_id is provided, only that opportunity is processed.
        When None, the entire bucket is scanned for raw files uploaded after
        since and all discovered opportunity IDs are processed. Opportunity IDs
        are merged from both raw and processed tiers so we reconcile documents
        even when an opportunity has only processed (all raw deleted).

        Returns:
            Tuple of (written_uris, deleted_uris).
        """
        if opportunity_id:
            logger.bind(opportunity_id=opportunity_id).info("GCS pipeline run started")
            return self.run_opportunity(opportunity_id, since=since)

        written: list[str] = []
        deleted: list[str] = []
        all_blobs = self._storage.list_blobs_since(prefix="", since=since)
        opp_ids_from_raw: set[str] = set()
        for blob_name in all_blobs:
            match = _RAW_BLOB_RE.match(blob_name)
            if match and self._get_handler(match.group(2)) is not None:
                opp_ids_from_raw.add(match.group(1))

        opp_ids_from_processed: set[str] = set()
        for opp_id, _source, _obj in self._storage.list_all_processed(
            opportunity_id=None, since=None
        ):
            opp_ids_from_processed.add(opp_id)

        all_opp_ids = opp_ids_from_raw | opp_ids_from_processed

        logger.info("GCS pipeline run started (all opportunities)", extra={})
        for opp_id in sorted(all_opp_ids):
            logger.bind(opportunity_id=opp_id).info("Processing opportunity")
            w, d = self.run_opportunity(opp_id, since=since)
            written.extend(w)
            deleted.extend(d)

        return written, deleted
