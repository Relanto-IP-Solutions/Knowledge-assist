"""Spreadsheet extraction for XLSX to per-sheet TXT outputs.

This extractor is intentionally self-contained:
- Input: XLSX bytes
- Output: one formatted TXT string per sheet (RAG-friendly)

It mirrors the local spreadsheet preprocessor's output shape:
SOURCE/FILE/SHEET/ROW_RANGE/COLUMNS/ROWS
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import load_workbook


if TYPE_CHECKING:  # pragma: no cover
    import pandas as pd


@dataclass(frozen=True)
class ExtractedSheetText:
    """One extracted sheet rendered to TXT plus its intended processed filename."""

    sheet_name: str
    processed_object_name: str
    text: str


def safe_sheet_name(sheet_name: str) -> str:
    """Convert an Excel sheet name into a Windows-safe filename component."""

    s = (sheet_name or "").lower().strip()
    s = s.replace(" ", "_")
    for ch in '<>:"/\\|?*':
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("._")


def processed_xlsx_sheet_filename(*, original_stem: str, sheet_name: str) -> str:
    """Match the local naming: _pre_xlsx_{safe_sheet}_{stem}.txt"""

    return f"_pre_xlsx_{safe_sheet_name(sheet_name)}_{original_stem}.txt"


def _is_blank(v: object) -> bool:
    # pandas is intentionally not imported at module import time (Cloud Run startup).
    import pandas as pd

    return bool(pd.isna(v) or (isinstance(v, str) and not v.strip()))


def _normalize_headers(columns: Iterable[object]) -> list[str]:
    """Normalize header cells without truncation; empty becomes '-'."""

    out: list[str] = []
    for col in columns:
        base = str(col).strip()
        if not base:
            out.append("-")
            continue
        tokens = base.split()
        out.append("_".join(tokens))
    return out


def _preprocess_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Fill empty cells for output stability (non-numeric -> '-', numeric -> 0)."""
    import pandas as pd

    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].fillna(0)
        else:
            df[col] = df[col].fillna("-")
    return df


def _format_sheet_txt(
    *,
    file_name: str,
    sheet_name: str,
    row_start: int,
    row_end: int,
    headers: list[str],
    data_df: pd.DataFrame,
) -> str:
    value_columns = [c for c in data_df.columns if c != "__row_number__"]
    if not headers:
        columns_line = " | ".join(["-"] * len(value_columns))
    else:
        columns_line = " | ".join(headers)

    lines: list[str] = [
        "SOURCE: spreadsheet",
        f"FILE: {file_name}",
        f"SHEET: {sheet_name}",
        f"ROW_RANGE: {row_start}-{row_end}",
        "",
        "COLUMNS:",
        columns_line,
        "",
        "ROWS:",
        "",
    ]

    def clean_cell(value: object) -> str:
        if value is None:
            return ""
        s = str(value)
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        s = " ".join(s.split())
        return s.strip()

    for _, row in data_df.iterrows():
        row_num = int(row["__row_number__"])
        vals = [clean_cell(row[c]) for c in value_columns]
        lines.append(f"Row {row_num}: " + " | ".join(vals))

    lines.append("")
    return "\n".join(lines)


class SpreadsheetExtractor:
    """Extract XLSX into one RAG-friendly TXT string per sheet."""

    def extract_sheets(
        self, content: bytes, object_name: str
    ) -> list[ExtractedSheetText]:
        # Import heavy deps lazily to keep Cloud Run cold-start fast.
        import pandas as pd

        wb = load_workbook(BytesIO(content), data_only=True)
        original_stem = object_name.rsplit("/", 1)[-1].rsplit(".", 1)[0]
        out: list[ExtractedSheetText] = []

        def _make_expanded_value(ws, merged_ranges):
            def expanded_value(row: int, col: int) -> object:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    return cell.value
                for mr in merged_ranges:
                    if (
                        mr.min_row <= row <= mr.max_row
                        and mr.min_col <= col <= mr.max_col
                    ):
                        return ws.cell(row=mr.min_row, column=mr.min_col).value
                return cell.value

            return expanded_value

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            merged_ranges = list(ws.merged_cells.ranges)
            expanded_value = _make_expanded_value(ws, merged_ranges)

            max_row = ws.max_row
            max_col = ws.max_column
            data: list[list[object]] = [
                [expanded_value(r, c) for c in range(1, max_col + 1)]
                for r in range(1, max_row + 1)
            ]
            raw_df = pd.DataFrame(data, dtype=object)

            headers, data_df = self._split_header_and_data(raw_df)
            if data_df.empty:
                # Still emit an empty-shell output? Current pipeline expects real content only.
                continue

            row_start = int(data_df["__row_number__"].min())
            row_end = int(data_df["__row_number__"].max())
            processed_name = processed_xlsx_sheet_filename(
                original_stem=original_stem,
                sheet_name=sheet_name,
            )
            out.append(
                ExtractedSheetText(
                    sheet_name=sheet_name,
                    processed_object_name=processed_name,
                    text=_format_sheet_txt(
                        file_name=object_name.rsplit("/", 1)[-1],
                        sheet_name=sheet_name,
                        row_start=row_start,
                        row_end=row_end,
                        headers=headers,
                        data_df=data_df,
                    ),
                )
            )

        wb.close()
        return out

    def _split_header_and_data(
        self, raw_df: pd.DataFrame
    ) -> tuple[list[str], pd.DataFrame]:
        """Heuristic header selection with banner/instruction reinsertion as extra rows."""
        import pandas as pd

        if raw_df.empty:
            return [], raw_df

        df = raw_df.dropna(axis=1, how="all").copy()
        if df.empty:
            return [], df

        def looks_like_banner_row(row: pd.Series) -> bool:
            vals = [v for v in row.tolist() if not _is_blank(v)]
            if len(vals) < 3:
                return False
            uniq = {str(v).strip() for v in vals}
            if len(uniq) != 1:
                return False
            return len(next(iter(uniq))) >= 40

        def looks_like_instruction_row(row: pd.Series) -> bool:
            vals = [str(v).strip() for v in row.tolist() if not _is_blank(v)]
            if len(vals) < 2:
                return False
            long_cells = [v for v in vals if len(v) >= 80]
            has_numbered_steps = any(
                ("1." in v or "2." in v or "3." in v) for v in vals
            )
            has_sentence_punct = any(
                (". " in v or "it's" in v.lower() or '"' in v) for v in vals
            )
            return bool(long_cells and (has_numbered_steps or has_sentence_punct))

        def header_score(row: pd.Series) -> int:
            vals = [str(v).strip() for v in row.tolist() if not _is_blank(v)]
            if not vals:
                return -10_000
            uniq = set(vals)
            score = 0
            score += min(len(vals), 10) * 2
            score += min(len(uniq), 10) * 3
            for v in vals:
                if len(v) >= 80:
                    score -= 10
                elif len(v) >= 40:
                    score -= 4
                else:
                    score += 1
                if any(ch in v for ch in [".", "•", "\n"]):
                    score -= 2
                if any(tok in v for tok in ["1.", "2.", "3.", "4."]):
                    score -= 3
            return score

        header_row_idx: int | None = None
        scan_limit = min(len(df), 12)
        min_header_score = 6
        for i in range(scan_limit):
            row = df.iloc[i]
            if row.isna().all():
                continue
            if looks_like_banner_row(row) or looks_like_instruction_row(row):
                continue
            non_empty = [v for v in row.tolist() if not _is_blank(v)]
            uniq = {str(v).strip() for v in non_empty if not _is_blank(v)}
            if len(non_empty) < 2 or len(uniq) < 2:
                continue
            if header_score(row) >= min_header_score:
                header_row_idx = i
                break

        if header_row_idx is None:
            first_row = df.iloc[0]
            use_second = bool(first_row.isna().all())
            header_row_idx = 1 if (use_second and len(df) > 1) else 0

        # Avoid direct pandas usage here (keeps imports lazy and prevents NameError).
        header_values = [
            ("" if _is_blank(v) else v) for v in df.iloc[header_row_idx].tolist()
        ]

        # Headerless detection (narrative blocks)
        header_non_empty = [str(v).strip() for v in header_values if str(v).strip()]
        headerless = False
        if len(header_non_empty) <= 2:
            lookahead = df.iloc[header_row_idx + 1 : header_row_idx + 4]
            for _, row in lookahead.iterrows():
                vals = row.tolist()
                filled = sum(0 if _is_blank(v) else 1 for v in vals)
                text_parts = [
                    str(v).strip() for v in vals if isinstance(v, str) and v.strip()
                ]
                if filled <= 2 and text_parts and sum(len(p) for p in text_parts) >= 60:
                    headerless = True
                    break

        if headerless:
            data_df = df.copy()
            data_df["__row_number__"] = data_df.index + 1
            data_df = self._trim_trailing_empty_rows(data_df)
            return [], _preprocess_dataframe(data_df)

        # Lock columns to header span
        non_empty_header_idx = [
            j for j, v in enumerate(header_values) if not _is_blank(v)
        ]
        if non_empty_header_idx:
            c_start = min(non_empty_header_idx)
            c_end = max(non_empty_header_idx) + 1
        else:
            c_start = 0
            c_end = df.shape[1]

        df_locked = df.iloc[:, c_start:c_end].copy()
        headers = _normalize_headers(header_values[c_start:c_end])

        # Extra info rows (banner/instruction above header) as first rows, merged text collapsed
        extra_rows_idx: list[int] = []
        for i in range(header_row_idx):
            row_full = df.iloc[i]
            if looks_like_banner_row(row_full) or looks_like_instruction_row(row_full):
                extra_rows_idx.append(i)

        def collapse_banner_row(row: pd.Series) -> pd.Series:
            row_vals = row.tolist()
            freq: dict[str, int] = {}
            for v in row_vals:
                if _is_blank(v):
                    continue
                sv = str(v).strip()
                if sv:
                    freq[sv] = freq.get(sv, 0) + 1

            banner_text: str | None = None
            repeated = [
                (sv, cnt) for sv, cnt in freq.items() if cnt >= 2 and len(sv) >= 40
            ]
            if repeated:
                banner_text = max(repeated, key=lambda x: x[1])[0]

            if banner_text is None:
                return row

            out_row = row.copy()
            first_used = False
            for col in out_row.index:
                cell = out_row[col]
                if _is_blank(cell):
                    continue
                if str(cell).strip() != banner_text:
                    continue
                if not first_used:
                    out_row[col] = f"(extra information) {banner_text}"
                    first_used = True
                else:
                    out_row[col] = None
            return out_row

        extra_rows: list[pd.Series] = [
            collapse_banner_row(df_locked.iloc[i]) for i in extra_rows_idx
        ]
        extra_df = (
            pd.DataFrame(extra_rows, columns=df_locked.columns)
            if extra_rows
            else pd.DataFrame(columns=df_locked.columns)
        )

        normal_df = df_locked.iloc[header_row_idx + 1 :].copy()
        if not extra_df.empty:
            extra_df["__row_number__"] = [i + 1 for i in extra_rows_idx]
            normal_df["__row_number__"] = normal_df.index + 1
            data_df = pd.concat([extra_df, normal_df], ignore_index=True)
        else:
            data_df = normal_df
            data_df["__row_number__"] = data_df.index + 1

        data_df = self._trim_trailing_empty_rows(data_df)
        data_df = _preprocess_dataframe(data_df)
        return headers, data_df

    def _trim_trailing_empty_rows(self, data_df: pd.DataFrame) -> pd.DataFrame:
        """Drop trailing rows that are truly empty before fillna turns them into '-'."""
        import pandas as pd

        value_cols = [c for c in data_df.columns if c != "__row_number__"]

        def is_blank_raw(v: object) -> bool:
            return bool(pd.isna(v) or (isinstance(v, str) and v.strip() == ""))

        flags: list[bool] = []
        for i in range(len(data_df)):
            vals = data_df.iloc[i][value_cols].tolist()
            flags.append(any(not is_blank_raw(v) for v in vals))

        if not any(flags):
            return data_df.iloc[0:0].copy()

        last = max(i for i, ok in enumerate(flags) if ok)
        return data_df.iloc[: last + 1].copy()
